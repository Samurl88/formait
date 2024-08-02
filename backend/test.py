from langchain.document_loaders import PyPDFLoader
from langchain.llms import OpenAI
from langchain.chains.question_answering import load_qa_chain
import xlrd  
import csv 
import pandas as pd 
import re
import json
json_regex = r'\{.*?\}'
import json
from pdf2image import convert_from_path
from openpyxl import load_workbook

def convertToTemplate():
    # read an excel file and convert  
    # into a dataframe object 
    workbook = load_workbook(filename="template.xlsx")
    sheet = workbook.active
    print(sheet.cell(row=8, column=2).value)
    replaceCoords = []
    for y in range(sheet.max_row):
        for x in range(sheet.max_column):
            if sheet.cell(row=y + 1, column=x + 1).value == "REPLACE":
                replaceCoords.append((y + 1, x + 1))
    print(replaceCoords)
    templateDictionary = {
        "date": (8, 5),
        "bill to": (10, 5),
        "ship to":(14, 5),
        "PO#": (20, 5),
        "production date": (21, 5),

    }
    
    
    #Locate your PDF here.


    pdf="./thing.pdf"
    #Load the PDF
    loader = PyPDFLoader(pdf)
    documents = loader.load()

    # api_key = "sk-?????"
    llm = OpenAI(model_name="gpt-4o")
    chain = load_qa_chain(llm,verbose=True)
    question = """With this pdf, find the date, billing address, and shipping address. Output your answer only in the following JSON format: 
    templateDictionary = {
        "date": string,
        "bill to": string (with line breaks),
        "ship to":string (with line breaks),
        "PO#": string,
        "production date": string,
        "expected date": string,
        "quantities": [
            quantity of cases of 8/64 oz Suntropics Mango Nectar,
            quantity of cases of 8/64 oz Suntropics Guava Nectar,
            quantity of cases of 8/64 oz  Suntropics Calamansi -,
            quantity of cases of 8/64 oz  Suntropics Passion OJ Guava 100% Juice
        ],
        "item numbers": [
            item number of 8/64 oz Suntropics Mango Nectar,
            item number of 8/64 oz Suntropics Guava Nectar,
            item number of 8/64 oz  Suntropics Calamansi -,
            item number of 8/64 oz  Suntropics Passion OJ Guava 100% Juice
        ],
        costs: [
            cost of 8/64 oz Suntropics Mango Nectar,
            cost of 8/64 oz Suntropics Guava Nectar,
            cost of 8/64 oz  Suntropics Calamansi -,
            cost of 8/64 oz  Suntropics Passion OJ Guava 100% Juice
        ],
        totalCosts: [
            total cost of 8/64 oz Suntropics Mango Nectar,
            total cost of 8/64 oz Suntropics Guava Nectar,
            total cost of 8/64 oz  Suntropics Calamansi -,
            total cost of 8/64 oz  Suntropics Passion OJ Guava 100% Juice
        ],
        totalQuantity: integer,
        netCost: integer in dollars
        
        "
    }"""

    response = str(chain.run(input_documents=documents, question=question))
    print(response)
    json_match = re.search(json_regex, response, re.DOTALL)
    responseJson = json.loads(json_match.group(0))
    print(responseJson)
    sheet.cell(row=8, column=5).value = responseJson["date"]
    #split addresses
    addresses = responseJson["bill to"].split("\n")
    for i in range(len(addresses)):   
        sheet.cell(row=10 + i, column=5).value = addresses[i]

    addresses = responseJson["ship to"].split("\n")
    for i in range(len(addresses)):   
        sheet.cell(row=14 + i, column=5).value = addresses[i]

    sheet.cell(row=20, column=5).value = responseJson["PO#"]
    sheet.cell(row=21, column=5).value = responseJson["production date"]
    sheet.cell(row=23, column=5).value = responseJson["expected date"]

    #quantities and stuff

    for i in range(4):
        sheet.cell(row=26+i, column=1).value = responseJson["quantities"][i]

    for i in range(4):
        sheet.cell(row=26+i, column=3).value = responseJson["item numbers"][i]
    
    for i in range(4):
        sheet.cell(row=26+i, column=6).value = responseJson["costs"][i]

    for i in range(4):
        sheet.cell(row=26+i, column=7).value = responseJson["totalCosts"][i]

    sheet.cell(row=30, column=1).value = responseJson["totalQuantity"]
    sheet.cell(row=30, column=7).value = responseJson["netCost"]
    workbook.save("test.xlsx")

convertToTemplate()