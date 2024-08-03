from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from pypdf import PdfReader
from openai import OpenAI
import time
import pandas as pd
import csv
import xlrd
from openpyxl import load_workbook
from pdf2image import convert_from_bytes
from io import BytesIO
import base64
import json
import re
from langchain.chains.question_answering import load_qa_chain
from langchain.document_loaders import PyPDFLoader
json_regex = r'\{.*?\}'

client = OpenAI()

app = Flask(__name__)
CORS(app)

# def sendToLLM2(pdf):
#     #Load the PDF
#     loader = PyPDFLoader(pdf)
#     documents = loader.load()

#     api_key = "sk-?????"
#     llm = OpenAI(openai_api_key=api_key)
#     chain = load_qa_chain(llm,verbose=True)
#     question = input("Enter your question here : ")
#     response = chain.run(input_documents=documents, question=question)
#     print(response) 
#     return response
# def sendToLLM(data):
#     completion = client.chat.completions.create(
#     model="gpt-4o-mini",
#     messages=[
#         {"role": "user", "content": data}
#     ]
#     )
#     return str(completion.choices[0].message.content)

def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

def convertTemplate3(bytes):
    workbook = load_workbook(filename="template.xlsx")
    sheet = workbook.active
    image = convert_from_bytes(bytes)[0]
    buffered = BytesIO()
    image.save("thing.jpeg")
    image_path = "thing.jpeg"
    question = """Let's think step by step. With this image, find the date, billing address, and shipping address. Output your answer only in the following JSON format: 
    templateDictionary = {
        "date": string,
        "bill to": string (with line breaks),
        "ship to":string (with line breaks),
        "PO#": string,
        "production date": string,
        "expected date": string,
        "quantities": [
            EA quantity of cases of 8/64 oz Suntropics Mango Nectar,
            EA quantity of cases of 8/64 oz Suntropics Guava Nectar,
            EA quantity of cases of 8/64 oz  Suntropics Calamansi -,
            EA quantity of cases of 8/64 oz  Suntropics Passion OJ Guava 100% Juice
        ],
        "item numbers": [
            MFG item number of Suntropics Mango Nectar,
            MFG item number of Suntropics Guava Nectar,
            MFG item number of Suntropics Calamansi -,
            MFG item number of Suntropics Passion OJ Guava 100% Juice
        ],
        costs: [
            cost of 8/64 oz Suntropics Mango Nectar,
            cost of 8/64 oz Suntropics Guava Nectar,
            cost of 8/64 oz  Suntropics Calamansi -,
            cost of 8/64 oz  Suntropics Passion OJ Guava 100% Juice
        ],
        totalCosts: [
            total/extended cost of 8/64 oz Suntropics Mango Nectar,
            total/extended cost of 8/64 oz Suntropics Guava Nectar,
            total/extended cost of 8/64 oz  Suntropics Calamansi -,
            total/extended cost of 8/64 oz  Suntropics Passion OJ Guava 100% Juice
        ],
        totalQuantity: integer, 
        
        "
    }"""
    # Getting the base64 string
    base64_image = encode_image(image_path)

    completion = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "user", "content": [
                {
                    "type": "text",
                    "text": question
                },
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{base64_image}"
                    }
                }
            ]
            }
        ]
    )

    response = str(completion.choices[0].message.content)
    print(response)
    json_match = re.search(json_regex, response, re.DOTALL)
    responseJson = json.loads(json_match.group(0))
    print(responseJson)
    sheet.cell(row=8, column=5).value = responseJson["date"]
    # split addresses
    addresses = responseJson["bill to"].split("\n")
    for i in range(len(addresses)):
        sheet.cell(row=10 + i, column=5).value = addresses[i]

    addresses = responseJson["ship to"].split("\n")
    for i in range(len(addresses)):
        sheet.cell(row=14 + i, column=5).value = addresses[i]

    sheet.cell(row=20, column=5).value = responseJson["PO#"]
    sheet.cell(row=21, column=5).value = responseJson["production date"]
    sheet.cell(row=23, column=5).value = responseJson["expected date"]

    # quantities and stuff
    totalQuantity = 0
    for i in range(4):
        try:
            totalQuantity += float(responseJson["quantities"][i])
        except Exception as e:
            print(e)
            pass
        sheet.cell(row=26+i, column=1).value = responseJson["quantities"][i]

    for i in range(4):
        sheet.cell(row=26+i, column=3).value = responseJson["item numbers"][i]

    for i in range(4):
        sheet.cell(row=26+i, column=6).value = responseJson["costs"][i]
    totalCost = 0
    for i in range(4):
        try:
            print(responseJson["totalCosts"][i])
            totalCost += float(responseJson["totalCosts"][i])
        except Exception as e:
            print(e)
            pass
        sheet.cell(row=26+i, column=7).value = responseJson["totalCosts"][i]

    sheet.cell(row=30, column=1).value = totalQuantity
    sheet.cell(row=30, column=7).value = totalCost
    workbook.title = "completed"
    workbook.save("test.xlsx")



@app.route("/", methods=['POST', "GET"])
def hello_world():
    pdf = request.files['pdf']
    template = request.files['template']

    # print(pdf.read())
    
    convertTemplate3(pdf.read())

    if pdf and template:
        return send_file('./test.xlsx', attachment_filename="completed.xlsx")
    else:
        return "no data"
    # return sendToLLM2("./thing.pdf")


app.run(debug=True)
